import os
import re
import shutil
import pyodbc
import openpyxl
from pdfminer.pdfparser import PDFParser, PDFDocument
from pdfminer.pdfinterp import PDFResourceManager, PDFPageInterpreter
from pdfminer.converter import PDFPageAggregator
from pdfminer.layout import LAParams, LTTextBox, LTTextLine
from dotenv import load_dotenv

# Load environment variables from .env file
load_dotenv()

REPORT_DIR = os.getenv("REPORT_DIR")
PROCESSED_DIR = os.getenv("PROCESSED_DIR")
EXCEL_FILE = os.getenv("EXCEL_FILE")

if not REPORT_DIR or not PROCESSED_DIR or not EXCEL_FILE:
    raise EnvironmentError("REPORT_DIR, PROCESSED_DIR, and EXCEL_FILE must be set in the .env file")

debug = False
excel_out = True

# Setup SQL connection
connprod = pyodbc.connect(
    'DRIVER={SQL Server};SERVER=PSGSQL;DATABASE=dbRedi;Trusted_Connection=yes;'
)
cursor = connprod.cursor()

if excel_out:
    wb = openpyxl.load_workbook(EXCEL_FILE)
    Data = wb["Data"]

# Process each PDF file in the report directory
files = os.listdir(REPORT_DIR)

for filename in files:
    if filename == 'Processed':
        print("All files processed")
        break

    if filename[6:9] == 'ARB':  # Skip ARB files, move to processed
        shutil.move(os.path.join(REPORT_DIR, filename), os.path.join(PROCESSED_DIR, filename))
        continue

    print(f"Processing file: {filename}")

    date_idx = filename.find('_sort')
    sub_date = filename[date_idx - 10:date_idx]
    sub_sort = filename[10:13]

    sort_map = {
        "Sun": "Sunrise",
        "Day": "Day",
        "Twi": "Twilight",
        "Nig": "Night",
    }
    sort = sort_map.get(sub_sort, "")

    with open(os.path.join(REPORT_DIR, filename), 'rb') as fp:
        parser = PDFParser(fp)
        doc = PDFDocument()
        parser.set_document(doc)
        doc.set_parser(parser)
        doc.initialize('')

        rsrcmgr = PDFResourceManager()
        laparams = LAParams()
        device = PDFPageAggregator(rsrcmgr, laparams=laparams)
        interpreter = PDFPageInterpreter(rsrcmgr, device)

        # Extract text from PDF pages to a temp txt file
        with open("Out_PDF.txt", "w", encoding='utf-8') as out_file:
            for pageNumber, page in enumerate(doc.get_pages()):
                if pageNumber == 8:
                    break
                interpreter.process_page(page)
                layout = device.get_result()
                for lt_obj in layout:
                    if isinstance(lt_obj, (LTTextBox, LTTextLine)):
                        out_file.write(lt_obj.get_text())

    # Initialize data containers
    Scan_Volume = []
    In_Volume = []
    Net_Srt_Volume = []
    Sort_Duration = []
    Off_End = []
    Package_Rejects = []
    ASX_Rejects = []
    Rejects = []
    No_ASX = []
    No_Tracking = []
    Label_Conflicts = []
    Oversized_Pkgs = []
    No_Decode = []
    Max_Rev_Rejects = []

    scanned = rej = ASX_rej = label = pkgs = 0

    with open("Out_PDF.txt", 'r', encoding='utf-8') as f:
        line_num = 0
        line = f.readline()
        while line:
            if line == "4 of 19":
                break
            line_num += 1
            if line.startswith('CACH'):
                if scanned < 4:
                    row = f.readline()
                    if re.search('[a-zA-Z]', row):
                        row = f.readline()
                        if re.search('[a-zA-Z]', row):
                            Scan_Volume.append(f.readline().rsplit(' ', 1)[0].rstrip())
                        else:
                            Scan_Volume.append(row.rsplit(' ', 1)[0].rstrip())
                    else:
                        Scan_Volume.append(row.rsplit(' ', 1)[0].rstrip())
                    scanned += 1
            elif line.startswith('Inducted Vol'):
                row = f.readline()
                if row.startswith('Net Sorted'):
                    Net_Srt_Volume.append(row.rsplit()[3].rstrip())
                    In_Volume.append('0')
                else:
                    In_Volume.append(row.rsplit()[0].rstrip())
            elif line.startswith('Net Sorted Vol'):
                Net_Srt_Volume.append(line.split()[3].rstrip())
            elif line.startswith('Sort Duration'):
                Sort_Duration.append(line.split()[3].rstrip())
            elif line.startswith('Package Rejects'):
                if line[15] == '\n':
                    pass
                elif rej % 2 == 0:
                    Package_Rejects.append(line.rsplit()[2].rstrip())
                    rej += 1
                else:
                    rej += 1
            elif line.startswith('ASX Rejects'):
                if len(line) > 13:
                    ASX_Rejects.append(line.rsplit()[2].rstrip())
            elif line.startswith('Oversized Pkgs'):
                if pkgs % 2 == 0:
                    row = f.readline()
                    if '%' not in row:
                        Oversized_Pkgs.append(row.split()[0].rstrip())
                pkgs += 1
            elif line.startswith('Max Rev Rejects'):
                No_Tracking.append(f.readline().rsplit(' ', 1)[0].rstrip())
                label_line = f.readline()
                if '%' not in label_line.rsplit(' ', 1)[0].rstrip():
                    Label_Conflicts.append(label_line.rsplit(' ', 1)[0].rstrip())
                else:
                    Label_Conflicts.append(f.readline().rsplit(' ', 1)[0].rstrip())

                decode_line = f.readline()
                if '%' not in decode_line.rsplit(' ', 1)[0].rstrip():
                    No_Decode.append(decode_line.rsplit(' ', 1)[0].rstrip())
                else:
                    No_Decode.append(f.readline().rsplit(' ', 1)[0].rstrip())
            elif line.startswith('Off the end'):
                Off_End.append(f.readline().rsplit(' ', 1)[0].rstrip())
            elif line.startswith('Rejects'):
                Rejects.append(f.readline().rsplit(' ', 1)[0].rstrip())

            line = f.readline()

    sort1 = [sort] * 4
    date1 = [sub_date] * 4
    Module = ['Primary', 'P41', 'P43', 'RC4']

    data = [
        Module,
        sort1,
        date1,
        Scan_Volume,
        In_Volume,
        Net_Srt_Volume,
        Sort_Duration,
        Off_End,
        Package_Rejects,
        ASX_Rejects,
        No_Tracking,
        Label_Conflicts,
        No_Decode,
        Oversized_Pkgs,
        Rejects
    ]

    # Pad data lists to length 4
    for item in data[3:]:
        while len(item) < 4:
            item.append('0')

    # Replace any alphabetic data with zero and convert numbers
    for item in data[3:]:
        for i in range(4):
            if re.search('[a-zA-Z]', item[i]):
                item[i] = '0'
            if '.' not in item[i]:
                item[i] = int(item[i].replace(',', ''))

    # Transpose data for inserting into Excel and DB
    data = list(map(list, zip(*data)))

    if excel_out:
        # Find first empty row in Excel sheet in column D (4)
        row = 1
        while Data.cell(row=row, column=4).value is not None:
            row += 1

        # Write data to Excel
        for row_data in data:
            for col_idx in range(1, 16):
                Data.cell(row=row, column=col_idx).value = row_data[col_idx - 1]
            row += 1

    # Insert data into SQL Server DB
    for row_data in data:
        try:
            cursor.executemany(
                "INSERT INTO tbl_New_Mod_Data "
                "(Module,Sort,SortDate,[Scan Volume],[Induct Volume],[Net Sorted Volume],[Sort Duration],[Off The End],"
                "[Package Rejects],[ASX Rejects],[No Tracking], [Label Conflicts], [No Decode], [Oversized Pkgs],[Total Rejects]) "
                "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
                [row_data],
            )
            cursor.commit()
        except Exception:
            print("Exception on insert, attempting duplicate DB")
            try:
                connprod_dup = pyodbc.connect('DRIVER={SQL Server};SERVER=PSGSQL;DATABASE=dbInfoXPIIBE;UID=wwUser;PWD=wwUser')
                cursor_dup = connprod_dup.cursor()
                cursor_dup.executemany(
                    "INSERT INTO tbl_New_Mod_Data_Duplicate "
                    "(Module,Sort,SortDate,[Scan Volume],[Induct Volume],[Net Sorted Volume],[Sort Duration],[Off The End],"
                    "[Package Rejects],[ASX Rejects],[No Tracking], [Label Conflicts], [No Decode], [Oversized Pkgs],[Total Rejects]) "
                    "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
                    [row_data],
                )
                cursor_dup.commit()
                cursor_dup.close()
                connprod_dup.close()
            except Exception as e:
                print(f"Duplicate DB insert also failed: {e}")

    # Move processed file to processed directory
    try:
        shutil.move(os.path.join(REPORT_DIR, filename), os.path.join(PROCESSED_DIR, filename))
    except Exception as e:
        print(f"Failed to move file {filename}: {e}")

    if debug:
        print("Scan Volume:", Scan_Volume)
        print("Inducted Volume:", In_Volume)
        print("Net Sorted Volume:", Net_Srt_Volume)
        print("No Tracking:", No_Tracking)
        print("Sort Duration:", Sort_Duration)
        print("Off The End:", Off_End)
        print("Package Rejects:", Package_Rejects)
        print("ASX Rejects:", ASX_Rejects)
        print("Label Conflicts:", Label_Conflicts)
        print("No Decode:", No_Decode)
        print("Oversized Packages:", Oversized_Pkgs)
        print("Total Rejects:", Rejects)
        print()

# Cleanup
cursor.close()
connprod.close()

if excel_out:
    wb.save(EXCEL_FILE)
